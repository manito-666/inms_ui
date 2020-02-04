# coding:utf-8
import os,sys,string
Path = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(Path)[0]
sys.path.append(rootPath)
from openpyxl import load_workbook
Path = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(Path)[0]
sys.path.append(rootPath)
import openpyxl,random
Path = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(Path)[0]
sys.path.append(rootPath)
from util.log.mylog import Log
from config.globalparam import data_path

def copy_excel(excelpath1,excelpath2):
    '''复制表格'''
    wb2=openpyxl.Workbook()
    wb2.save(excelpath2)

    #读取数据
    wb1=openpyxl.load_workbook(excelpath1)
    wb2=openpyxl.load_workbook(excelpath2)
    sheetnames=wb1.sheetnames
    sheets1=wb1.sheetnames
    sheets2=wb2.sheetnames
    sheet1=wb1[sheets1[0]]
    sheet2=wb2[sheets2[0]]
    max_row=sheet1.max_row
    max_column=sheet1.max_column


    for m in list(range(1, max_row + 1)):
        for n in list(range(97, 97 + max_column)):  # chr(97)='a'
            n = chr(n)  # ASCII字符
            i = '%s%d' % (n, m)  # 单元格编号
            cell1 = sheet1[i].value # 获取data单元格数据
            sheet2[i].value = cell1  # 赋值到test单元格
    wb2.save(excelpath2)  # 保存数据
    wb1.close()  # 关闭excel
    wb2.close()

class Write_excel(object):
    '''修改excel数据'''
    def __init__(self, filename,):
        self.filename=filename  #定义文件名获取方法
        self.wb = load_workbook(self.filename)  #打开已有的工作簿
        self.ws = self.wb.active  # 激活sheet,得到活动表表名

    def write(self,i,row_n, col_n, value):
        '''写入数据，如(0,2,3，"hello"),在下标为0的表名第二行第三列写入数据"hello"'''
        #得到工作簿的所有工作表
        self.sheetnames = self.wb.sheetnames
        #得到工作簿的第i个下标的sheetname
        self.ws = self.wb[self.sheetnames[i]]
        #通过行列读
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)
        Log().info('更新数据完成！')


    def printing(self):
        '''打印data表格'''
        wb=openpyxl.load_workbook(data_path)
        sheets=wb.sheetnames

        for i in range(len(sheets)):
            sheet = wb[sheets[i]]
            print('\n\n第' + str(i + 1) + '个sheet: ' + sheet.title + '->>>')
            for r in range(1, sheet.max_row + 1):
                if r == 1:
                    print('\n' + ''.join(
                        [str(sheet.cell(row=r, column=c).value).ljust(17) for c in range(1, sheet.max_column + 1)]))
                else:
                    print(''.join([str(sheet.cell(row=r, column=c).value).ljust(20) for c in range(1, sheet.max_column + 1)]))




if __name__ == '__main__':
    wb = load_workbook(data_path)
    sheetnames=wb.sheetnames
    print(sheetnames)
    set=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z']
    b='测试'+random.choice(set)
    c="租户"+random.choice(set)
    d="专网等级模版"+random.choice(set)
    e="切片模版"+random.choice(set)
    phone_a="185"
    phone_b=random.randint(10000000,99999999)
    phone = phone_a + str(phone_b)
    username=[]
    for i in range(8):
        if random.randint(0, 1):
            letter = random.choice(string.ascii_letters)
            username.append(letter)
        else:
            letter = random.choice(string.digits)
            username.append(letter)
    f=("".join(username))
    Write_excel(data_path).write(0,2,1,format(b))
    # Write_excel(data_path).write(1, 2, 1, b)
    #
    # Write_excel(data_path).write(2, 2, 1,c )
    # Write_excel(data_path).write(2, 2, 3,phone)
    #
    # Write_excel(data_path).write(3, 2, 1, phone)
    # Write_excel(data_path).write(4, 2, 1, d)
    # Write_excel(data_path).write(5, 2, 1, e)
    # Write_excel(data_path).write(6, 2, 1, f)

